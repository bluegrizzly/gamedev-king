import io
import os
import re
from pathlib import Path
from typing import List, Literal, Optional
from uuid import UUID

from fastapi import File, Form, HTTPException, UploadFile
from openai import OpenAI
from pydantic import BaseModel, Field
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from supabase import Client, create_client

from local_paths import get_local_project_path, require_local_project_path

EMBEDDING_MODEL = "text-embedding-3-small"
MAX_TOP_K = 20
DEFAULT_TOP_K = 12
EMBEDDING_BATCH_SIZE = 50
VALID_SCOPES = ("generic", "project", "hybrid")


class RetrieveRequest(BaseModel):
    query: str = Field(min_length=1)
    top_k: int = Field(default=DEFAULT_TOP_K, ge=1, le=MAX_TOP_K)
    source_id: Optional[UUID] = None
    agent_id: Optional[str] = None
    agent_ids: Optional[List[str]] = None
    scope: Literal["generic", "project", "hybrid"] = "hybrid"
    project_key: Optional[str] = None


class RetrieveResult(BaseModel):
    source_id: UUID
    chunk_index: int
    content: str
    score: float
    title: str
    scope: str
    project_key: Optional[str]


class RetrieveResponse(BaseModel):
    query: str
    results: List[RetrieveResult]


def retrieve_chunks(
    query: str,
    top_k: int,
    source_id: Optional[UUID],
    agent_ids: Optional[List[str]],
    scope: str,
    project_key: Optional[str],
) -> List[RetrieveResult]:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Missing OPENAI_API_KEY.")

    openai_client = OpenAI(api_key=api_key)
    embedding_response = openai_client.embeddings.create(
        model=EMBEDDING_MODEL,
        input=[query],
    )
    query_embedding = embedding_response.data[0].embedding

    supabase = get_supabase_client()
    rpc_payload = {
        "query_embedding": query_embedding,
        "match_count": top_k,
        "source_filter": str(source_id) if source_id else None,
        "agent_filter": agent_ids,
        "scope_mode": scope,
        "project_key_filter": project_key,
    }
    result = supabase.rpc("match_chunks", rpc_payload).execute()
    rows = result.data or []
    return [
        RetrieveResult(
            source_id=UUID(row["source_id"]),
            chunk_index=row["chunk_index"],
            content=row["content"],
            score=float(row["distance"]),
            title=row["title"],
            scope=row.get("scope", "generic"),
            project_key=row.get("project_key"),
        )
        for row in rows
    ]


def chunk_text(text: str, chunk_size: int = 1200, overlap: int = 200) -> List[str]:
    if chunk_size <= overlap:
        raise ValueError("chunk_size must be greater than overlap")
    cleaned = text.replace("\x00", " ").strip()
    if not cleaned:
        return []
    chunks: list[str] = []
    step = chunk_size - overlap
    start = 0
    while start < len(cleaned):
        end = min(start + chunk_size, len(cleaned))
        chunk = cleaned[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start += step
    return chunks


def extract_pdf_text(file_bytes: bytes) -> str:
    reader = PdfReader(io.BytesIO(file_bytes))
    pages_lines: list[list[str]] = []
    header_counts: dict[str, int] = {}
    footer_counts: dict[str, int] = {}

    for page in reader.pages:
        extracted = page.extract_text() or ""
        lines = [line.strip() for line in extracted.splitlines() if line.strip()]
        if not lines:
            continue
        # Skip very short pages (likely blank or separator pages)
        if len(" ".join(lines)) < 50:
            continue
        pages_lines.append(lines)
        header_counts[lines[0]] = header_counts.get(lines[0], 0) + 1
        footer_counts[lines[-1]] = footer_counts.get(lines[-1], 0) + 1

    header_blacklist = {line for line, count in header_counts.items() if count >= 3}
    footer_blacklist = {line for line, count in footer_counts.items() if count >= 3}

    cleaned_pages: list[str] = []
    for lines in pages_lines:
        filtered: list[str] = []
        for idx, line in enumerate(lines):
            if idx == 0 and line in header_blacklist:
                continue
            if idx == len(lines) - 1 and line in footer_blacklist:
                continue
            if line.isdigit():
                continue
            filtered.append(line)
        if filtered:
            paragraph = " ".join(filtered)
            paragraph = re.sub(r"\s+", " ", paragraph).strip()
            if paragraph:
                cleaned_pages.append(paragraph)

    return "\n\n".join(cleaned_pages).strip()


def get_supabase_client() -> Client:
    supabase_url = os.getenv("SUPABASE_URL")
    supabase_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not supabase_key:
        raise RuntimeError("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")
    return create_client(supabase_url, supabase_key)


def project_exists(supabase: Client, project_key: str) -> bool:
    if not project_key:
        return False
    result = (
        supabase.table("projects")
        .select("project_key")
        .eq("project_key", project_key)
        .limit(1)
        .execute()
    )
    return bool(result.data)


def get_default_project_key(supabase: Client) -> Optional[str]:
    result = (
        supabase.table("projects")
        .select("project_key")
        .order("created_at", desc=False)
        .limit(1)
        .execute()
    )
    if result.data:
        return result.data[0].get("project_key")
    return None


def get_default_project_key_value() -> Optional[str]:
    supabase = get_supabase_client()
    return get_default_project_key(supabase)


def get_project_display_name(supabase: Client, project_key: str) -> Optional[str]:
    if not project_key:
        return None
    result = (
        supabase.table("projects")
        .select("display_name")
        .eq("project_key", project_key)
        .limit(1)
        .execute()
    )
    if result.data:
        return result.data[0].get("display_name")
    return None


def resolve_project_path(project_key: Optional[str]) -> Optional[str]:
    cleaned_key = project_key.strip() if project_key else ""
    if cleaned_key:
        path = get_local_project_path(cleaned_key)
        if path:
            return path
    supabase = get_supabase_client()
    default_key = get_default_project_key(supabase)
    if not default_key:
        return None
    return get_local_project_path(default_key)


def require_project_path(project_key: Optional[str]) -> str:
    return require_local_project_path(project_key)


def resolve_scope_and_project_key(scope: str, project_key: Optional[str]) -> tuple[str, Optional[str]]:
    cleaned_scope = (scope or "hybrid").strip().lower()
    if cleaned_scope not in VALID_SCOPES:
        raise HTTPException(status_code=400, detail="Invalid scope.")
    cleaned_project_key = project_key.strip() if project_key else ""
    if cleaned_scope in ("project", "hybrid") and not cleaned_project_key:
        supabase = get_supabase_client()
        default_key = get_default_project_key(supabase)
        if default_key:
            return cleaned_scope, default_key
        return "generic", None
    if cleaned_scope == "generic":
        return "generic", None
    return cleaned_scope, cleaned_project_key


def extract_docx_text(file_bytes: bytes) -> str:
    document = Document(io.BytesIO(file_bytes))
    paragraphs = [para.text.strip() for para in document.paragraphs if para.text.strip()]
    return "\n".join(paragraphs).strip()


def extract_xlsx_text(file_bytes: bytes) -> str:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in (row or [])]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append("\n".join(rows))
    wb.close()
    text = "\n\n".join(parts).strip()
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def extract_text_from_file(filename: str, file_bytes: bytes) -> str:
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return extract_pdf_text(file_bytes)
    if lower.endswith(".docx"):
        return extract_docx_text(file_bytes)
    if lower.endswith(".xlsx"):
        return extract_xlsx_text(file_bytes)
    raise HTTPException(status_code=400, detail="Unsupported file type.")


def _allowed_upload_extensions() -> tuple[str, ...]:
    return (".pdf", ".docx", ".xlsx")


def upload_pdf(
    file: UploadFile = File(...),
    title: Optional[str] = Form(None),
    agent_ids: Optional[List[str]] = Form(None),
    scope: str = Form("generic"),
    project_key: Optional[str] = Form(None),
    source_path: Optional[str] = Form(None),
) -> dict:
    if not file.filename:
        raise HTTPException(status_code=400, detail="A file is required.")
    filename_lower = file.filename.lower()
    allowed = _allowed_upload_extensions()
    if not any(filename_lower.endswith(ext) for ext in allowed):
        raise HTTPException(
            status_code=400,
            detail=f"Only {', '.join(allowed)} files are supported.",
        )
    if file.content_type not in (
        None,
        "",
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ):
        raise HTTPException(status_code=400, detail="Invalid content type.")

    file_bytes = file.file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Empty upload.")

    text = extract_text_from_file(file.filename, file_bytes)
    if not text.strip():
        raise HTTPException(status_code=400, detail="No text extracted from file.")

    chunks = chunk_text(text, chunk_size=1200, overlap=200)
    if not chunks:
        raise HTTPException(status_code=400, detail="No chunks produced.")

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="Missing OPENAI_API_KEY.")

    openai_client = OpenAI(api_key=api_key)
    try:
        embeddings: list[list[float]] = []
        for i in range(0, len(chunks), EMBEDDING_BATCH_SIZE):
            batch = chunks[i : i + EMBEDDING_BATCH_SIZE]
            embedding_response = openai_client.embeddings.create(
                model=EMBEDDING_MODEL,
                input=batch,
            )
            embeddings.extend([item.embedding for item in embedding_response.data])
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Embedding failed: {exc}") from exc

    if len(embeddings) != len(chunks):
        raise HTTPException(status_code=502, detail="Embedding count mismatch.")

    source_title = title.strip() if title and title.strip() else file.filename
    cleaned_agents = [agent.strip() for agent in (agent_ids or []) if agent.strip()]
    source_agents = cleaned_agents if cleaned_agents else None
    cleaned_scope = scope.strip().lower()
    if cleaned_scope not in ("generic", "project"):
        raise HTTPException(status_code=400, detail="Invalid scope.")
    cleaned_project_key = project_key.strip() if project_key else ""
    if cleaned_scope == "project":
        if not cleaned_project_key:
            raise HTTPException(status_code=400, detail="project_key is required for project scope.")
    else:
        cleaned_project_key = ""

    try:
        supabase = get_supabase_client()
        if cleaned_scope == "project" and not project_exists(supabase, cleaned_project_key):
            raise HTTPException(status_code=400, detail="Unknown project_key.")
        source_payload = {
            "title": source_title,
            "agent_ids": source_agents,
            "scope": cleaned_scope,
            "project_key": cleaned_project_key or None,
            "source_path": source_path.strip() if source_path and source_path.strip() else None,
        }
        source_result = supabase.table("sources").insert(source_payload).execute()
        if not source_result.data:
            raise RuntimeError("No source row returned.")
        source_id = source_result.data[0]["id"]

        chunk_rows = [
            {
                "source_id": source_id,
                "chunk_index": idx,
                "content": chunk,
                "embedding": embeddings[idx],
                "scope": cleaned_scope,
                "project_key": cleaned_project_key or None,
            }
            for idx, chunk in enumerate(chunks)
        ]
        supabase.table("chunks").insert(chunk_rows).execute()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Supabase insert failed: {exc}") from exc

    return {"source_id": source_id, "chunks_indexed": len(chunks)}


def retrieve(body: RetrieveRequest) -> RetrieveResponse:
    query = body.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query must be non-empty.")

    try:
        agent_filter = body.agent_ids
        if not agent_filter and body.agent_id:
            agent_filter = [body.agent_id]
        cleaned_scope, cleaned_project_key = resolve_scope_and_project_key(
            body.scope,
            body.project_key,
        )
        results = retrieve_chunks(
            query,
            body.top_k,
            body.source_id,
            agent_filter,
            cleaned_scope,
            cleaned_project_key or None,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Vector search failed: {exc}") from exc
    return RetrieveResponse(query=query, results=results)


def list_sources() -> list[dict]:
    try:
        supabase = get_supabase_client()
        result = (
            supabase.table("sources")
            .select("id,title,created_at,agent_id,agent_ids,scope,project_key,source_path")
            .order("created_at", desc=True)
            .execute()
        )
        return result.data or []
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to load sources: {exc}") from exc


def delete_source(source_id: UUID) -> dict:
    try:
        supabase = get_supabase_client()
        result = supabase.table("sources").delete().eq("id", str(source_id)).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Source not found.")
        return {"deleted": True, "source_id": str(source_id)}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to delete source: {exc}") from exc


def refresh_source(source_id: UUID) -> dict:
    try:
        supabase = get_supabase_client()
        source_result = (
            supabase.table("sources")
            .select("id,title,scope,project_key,agent_id,agent_ids,source_path")
            .eq("id", str(source_id))
            .limit(1)
            .execute()
        )
        if not source_result.data:
            raise HTTPException(status_code=404, detail="Source not found.")
        source = source_result.data[0]
        source_path = source.get("source_path") or ""
        if not source_path:
            raise HTTPException(status_code=400, detail="source_path is missing for this source.")
        if not os.path.exists(source_path):
            raise HTTPException(status_code=404, detail="source_path does not exist.")

        file_bytes = Path(source_path).read_bytes()
        if not file_bytes:
            raise HTTPException(status_code=400, detail="Source file is empty.")

        text = extract_text_from_file(source_path, file_bytes)
        if not text.strip():
            raise HTTPException(status_code=400, detail="No text extracted from source file.")

        chunks = chunk_text(text, chunk_size=1200, overlap=200)
        if not chunks:
            raise HTTPException(status_code=400, detail="No chunks produced.")

        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Missing OPENAI_API_KEY.")

        openai_client = OpenAI(api_key=api_key)
        embeddings: list[list[float]] = []
        for i in range(0, len(chunks), EMBEDDING_BATCH_SIZE):
            batch = chunks[i : i + EMBEDDING_BATCH_SIZE]
            embedding_response = openai_client.embeddings.create(
                model=EMBEDDING_MODEL,
                input=batch,
            )
            embeddings.extend([item.embedding for item in embedding_response.data])

        if len(embeddings) != len(chunks):
            raise HTTPException(status_code=502, detail="Embedding count mismatch.")

        supabase.table("chunks").delete().eq("source_id", str(source_id)).execute()
        chunk_rows = [
            {
                "source_id": str(source_id),
                "chunk_index": idx,
                "content": chunk,
                "embedding": embeddings[idx],
                "scope": source.get("scope") or "generic",
                "project_key": source.get("project_key"),
            }
            for idx, chunk in enumerate(chunks)
        ]
        supabase.table("chunks").insert(chunk_rows).execute()
        return {"updated": True, "source_id": str(source_id), "chunks_indexed": len(chunks)}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to refresh source: {exc}") from exc
